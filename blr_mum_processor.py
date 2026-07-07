"""
Bengaluru & Mumbai XD Allocation Processor
-------------------------------------------
Skeletal script designed to handle specialized multi-sheet allocations
and calculations for Bengaluru and Mumbai cities.

Usage:
    python blr_mum_processor.py run --date 2026-06-23 --city bengaluru
    python blr_mum_processor.py run --date 2026-06-23 --city mumbai
    python blr_mum_processor.py serve --port 5060
"""

import argparse
import os
import sys
import json
from datetime import datetime, date, timedelta
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ─────────────────────────────────────────────
# CONFIGURATION
# ─────────────────────────────────────────────
SERVICE_ACCOUNT_KEY = "xd-allocation-9640b0ce66d2.json"
OUTPUT_DIR = "outputs"

CITIES = {
    "bengaluru": {
        "label": "Bengaluru",
        "tab_prefix": "BLR",
        "sheet_id": "YOUR_BENGALURU_SHEET_ID_HERE",
        "sheets_required": ["PO", "Indent Plan", "Store Master"],  # Placeholder list of sheets
    },
    "mumbai": {
        "label": "Mumbai",
        "tab_prefix": "MUM",
        "sheet_id": "YOUR_MUMBAI_SHEET_ID_HERE",
        "sheets_required": ["PO", "Indent Plan", "Route Matrix"],  # Placeholder list of sheets
    }
}

# Styling colors
CLR = {
    "header_bg": "1F4E79",   # Dark Blue
    "header_fg": "FFFFFF",   # White
    "section_bg": "BDD7EE",  # Light Blue
    "alt_row": "EBF3FB",     # Soft Alternate Row tint
    "total_bg": "FFE699",    # Yellow
    "highlight_bg": "FFF2CC" # Soft Yellow / Cream
}

border_side = Side(style="medium", color="FFCCCCCC")
BORDER = Border(left=border_side, right=border_side, top=border_side, bottom=border_side)

# ─────────────────────────────────────────────
# GOOGLE SHEETS LOADING
# ─────────────────────────────────────────────
def fetch_worksheet_data(sheet_id: str, tab_name: str) -> pd.DataFrame:
    """
    Fetches raw tab data from the given spreadsheet and tab name.
    """
    if "YOUR_" in sheet_id:
        raise RuntimeError(f"Spreadsheet ID is not configured yet.")

    import gspread
    from google.oauth2.service_account import Credentials

    scopes = [
        "https://www.googleapis.com/auth/spreadsheets.readonly",
        "https://www.googleapis.com/auth/drive.readonly",
    ]
    creds = Credentials.from_service_account_file(SERVICE_ACCOUNT_KEY, scopes=scopes)
    gc = gspread.authorize(creds)
    sh = gc.open_by_key(sheet_id)
    ws = sh.worksheet(tab_name)
    all_values = ws.get_all_values()
    
    if not all_values:
        return pd.DataFrame()
        
    headers = [h.strip() for h in all_values[0]]
    df = pd.DataFrame(all_values[1:], columns=headers)
    
    # Strip whitespaces from string values
    for c in df.columns:
        df[c] = df[c].astype(str).str.strip()
        
    return df

# ─────────────────────────────────────────────
# SKELETAL PROCESSING & CALCULATION FUNCTIONS
# ─────────────────────────────────────────────
def process_bengaluru_data(target_date: date, city_cfg: dict) -> dict:
    """
    Placeholder for Bengaluru's calculations and multi-sheet joins.
    Returns a dict of DataFrames mapped to sheet names.
    """
    print(f"[BLR] Processing Bengaluru multi-sheet calculations...")
    
    # Example raw tab names:
    # po_tab = f"{city_cfg['tab_prefix']}_{target_date.strftime('%Y-%m-%d')}_PO"
    
    # TODO: Fetch sheets e.g.:
    # po_df = fetch_worksheet_data(city_cfg['sheet_id'], po_tab)
    # indent_plan_df = fetch_worksheet_data(city_cfg['sheet_id'], 'BLR Indent Plan')
    
    # TODO: Run calculations, joins, and vertical-specific modifications here.
    
    po_cols = [
        "City", "Store Site ID", "FSN", "QTY", "Store ID", "SLA",
        "Supplier ID", "Contract ID", "Po No", "Title", "Brand",
        "Vertical", "Chiller/Non chiller Tag"
    ]
    
    # Return dummy dataframes for compilation
    sheets_out = {
        "PO": pd.DataFrame(columns=po_cols),
        "Calculations": pd.DataFrame({"Message": ["Multi-sheet formulas placeholder"]})
    }
    return sheets_out

def process_mumbai_data(target_date: date, city_cfg: dict) -> dict:
    """
    Placeholder for Mumbai's calculations and multi-sheet joins.
    Returns a dict of DataFrames mapped to sheet names.
    """
    print(f"[MUM] Processing Mumbai multi-sheet calculations...")
    
    po_cols = [
        "City", "Store Site ID", "FSN", "QTY", "Store ID", "SLA",
        "Supplier ID", "Contract ID", "Po No", "Title", "Brand",
        "Vertical", "Chiller/Non chiller Tag"
    ]
    
    # Return dummy dataframes for compilation
    sheets_out = {
        "PO": pd.DataFrame(columns=po_cols),
        "Calculations": pd.DataFrame({"Message": ["Multi-sheet formulas placeholder"]})
    }
    return sheets_out

# ─────────────────────────────────────────────
# EXCEL GENERATION & STYLING HELPERS
# ─────────────────────────────────────────────
def _hdr_style(cell):
    cell.font = Font(bold=True, color=CLR["header_fg"], name="Arial", size=10)
    cell.fill = PatternFill("solid", start_color=CLR["header_bg"])
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = BORDER

def _data_style(cell, bold=False, bg=None, fg="000000", num_fmt=None, align="center"):
    cell.font = Font(bold=bold, color=fg, name="Arial", size=10)
    cell.alignment = Alignment(horizontal=align, vertical="center")
    cell.border = BORDER
    if bg:
        cell.fill = PatternFill("solid", start_color=bg)
    if num_fmt:
        cell.number_format = num_fmt

def _auto_width(ws, min_w=10, max_w=40):
    for col in ws.columns:
        col_letter = get_column_letter(col[0].column)
        length = 0
        for c in col:
            val = str(c.value or "")
            val_len = 10 if val.startswith("=") else len(val)
            if val_len > length:
                length = val_len
        ws.column_dimensions[col_letter].width = min(max(length + 3, min_w), max_w)

def _write_df_to_sheet(ws, df: pd.DataFrame):
    # Header
    for col_idx, col_name in enumerate(df.columns, 1):
        _hdr_style(ws.cell(row=1, column=col_idx, value=col_name))
    ws.row_dimensions[1].height = 20

    # Data
    for row_idx, row in enumerate(df.itertuples(index=False), 2):
        is_alt = (row_idx % 2 == 0)
        for col_idx, value in enumerate(row, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            
            # Format numbers vs text vs dates
            if isinstance(value, (int, float)):
                _data_style(cell, bg=CLR["alt_row"] if is_alt else None, align="right", num_fmt="#,##0")
            else:
                _data_style(cell, bg=CLR["alt_row"] if is_alt else None, align="center")

# ─────────────────────────────────────────────
# MAIN BUILDER
# ─────────────────────────────────────────────
def build_report(target_date: date, output_path: str, city_key: str) -> str:
    city_cfg = CITIES[city_key]
    
    # Step 1: Run calculations
    if city_key == "bengaluru":
        sheets_data = process_bengaluru_data(target_date, city_cfg)
    elif city_key == "mumbai":
        sheets_data = process_mumbai_data(target_date, city_cfg)
    else:
        raise ValueError(f"Unknown city key '{city_key}'")

    # Step 2: Create workbook and populate sheets
    wb = Workbook()
    wb.remove(wb.active)  # Remove default active sheet

    for sheet_name, df in sheets_data.items():
        ws = wb.create_sheet(sheet_name)
        _write_df_to_sheet(ws, df)
        ws.freeze_panes = "A2"
        _auto_width(ws)

    os.makedirs(os.path.dirname(output_path) if os.path.dirname(output_path) else ".", exist_ok=True)
    wb.save(output_path)
    print(f"[{city_cfg['label']}] Report saved -> {output_path}")
    return output_path

# ─────────────────────────────────────────────
# FLASK SERVER
# ─────────────────────────────────────────────
def create_app():
    from flask import Flask, request, jsonify, send_file
    import io, zipfile

    app = Flask(__name__)

    @app.route("/health", methods=["GET"])
    def health():
        return jsonify({"status": "ok", "cities": list(CITIES.keys())})

    @app.route("/process", methods=["POST"])
    def process():
        body = request.get_json(force=True, silent=True) or {}
        date_str = body.get("date") or request.args.get("date")
        city_key = (body.get("city") or request.args.get("city", "all")).lower()

        if not date_str:
            return jsonify({"error": "Missing 'date' field. Send { \"date\": \"YYYY-MM-DD\" }"}), 400

        try:
            target_date = datetime.strptime(date_str, "%Y-%m-%d").date()
        except ValueError:
            return jsonify({"error": "Invalid date format. Use YYYY-MM-DD"}), 400

        delivery_date = target_date + timedelta(days=1)
        delivery_str = delivery_date.strftime('%Y%m%d')

        cities_to_run = list(CITIES.keys()) if city_key == "all" else [city_key]
        for ck in cities_to_run:
            if ck not in CITIES:
                return jsonify({"error": f"Unknown city '{ck}'"}), 400

        if len(cities_to_run) == 1:
            ck = cities_to_run[0]
            label = CITIES[ck]["label"]
            fname = f"BLR_MUM_Allocation_{label}_{delivery_str}Delivery.xlsx"
            opath = os.path.join(OUTPUT_DIR, fname)
            try:
                build_report(target_date, opath, ck)
            except Exception as e:
                return jsonify({"error": str(e)}), 500
            return send_file(opath, as_attachment=True, download_name=fname)
        else:
            zip_buf = io.BytesIO()
            with zipfile.ZipFile(zip_buf, "w", zipfile.ZIP_DEFLATED) as zf:
                for ck in cities_to_run:
                    label = CITIES[ck]["label"]
                    fname = f"BLR_MUM_Allocation_{label}_{delivery_str}Delivery.xlsx"
                    opath = os.path.join(OUTPUT_DIR, fname)
                    try:
                        build_report(target_date, opath, ck)
                        zf.write(opath, fname)
                    except Exception as e:
                        print(f"[SKIP] Failed {label}: {e}")
            zip_buf.seek(0)
            zip_name = f"BLR_MUM_Allocation_All_{delivery_str}Delivery.zip"
            return send_file(zip_buf, as_attachment=True, download_name=zip_name)

    return app

# ─────────────────────────────────────────────
# CLI ENTRY POINT
# ─────────────────────────────────────────────
if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Bengaluru & Mumbai XD Allocation Processor")
    subparsers = parser.add_subparsers(dest="cmd")

    run_parser = subparsers.add_parser("run", help="Run processor for a specific date")
    run_parser.add_argument("--date", required=True, help="Date in YYYY-MM-DD format")
    run_parser.add_argument("--city", default="all", choices=list(CITIES.keys()) + ["all"])
    run_parser.add_argument("--out", default=None, help="Output path (for single city)")

    serve_parser = subparsers.add_parser("serve", help="Start Flask API server")
    serve_parser.add_argument("--port", default=5060, type=int)

    args = parser.parse_args()

    if args.cmd == "run":
        target_date = datetime.strptime(args.date, "%Y-%m-%d").date()
        delivery_date = target_date + timedelta(days=1)
        delivery_str = delivery_date.strftime('%Y%m%d')

        cities_to_run = list(CITIES.keys()) if args.city == "all" else [args.city]
        for city_key in cities_to_run:
            label = CITIES[city_key]["label"]
            if args.out and len(cities_to_run) == 1:
                out = args.out
            else:
                out = os.path.join(OUTPUT_DIR, f"BLR_MUM_Allocation_{label}_{delivery_str}Delivery.xlsx")
            try:
                build_report(target_date, out, city_key)
            except RuntimeError as e:
                print(f"[SKIP] {label}: {e}")

    elif args.cmd == "serve":
        app = create_app()
        print(f"[OK] BLR/MUM Processor API running on port {args.port}")
        app.run(host="0.0.0.0", port=args.port, debug=False)
    else:
        parser.print_help()
