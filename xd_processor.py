"""
XD Allocation Processor – Multi-City (Trichy / Chennai / Coimbatore)
---------------------------------------------------------------------
Fetches the Indent Google Sheet for a given date and city,
merges with the XD Allocation PO data, builds pivot tables,
cross-verifies totals, and outputs a formatted Excel report.

Usage (standalone):
    python xd_processor.py run --date 2026-06-23              # all cities
    python xd_processor.py run --date 2026-06-23 --city trichy
    python xd_processor.py run --date 2026-06-23 --city chennai

Usage (via Flask API):
    POST /process  { "date": "2026-06-23" }                   # all cities -> ZIP
    POST /process  { "date": "2026-06-23", "city": "trichy" } # single city -> xlsx
"""

import argparse
import os
import sys
import json
from datetime import datetime, date, timedelta

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows

# ─────────────────────────────────────────────
# CONFIG  – fill these in before running
# ─────────────────────────────────────────────

# Service-account key shared across all cities
SERVICE_ACCOUNT_KEY = "xd-allocation-9640b0ce66d2.json"

# Per-city config
# tab_prefix : the letters before _YYYY-MM-DD_PO in the Google Sheet tab name
# sheet_id   : ID from the spreadsheet URL  .../spreadsheets/d/<ID>/edit
CITIES = {
    "trichy": {
        "label":           "Trichy",
        "tab_prefix":      "Tri",
        "sheet_id":        "1lXWT-9x3WbNpYYXLZcP4_DEwfuAl1VWbWpouNfTzwXU", #main sheet link updated
        "indent_plan_tab": "Trichy Indent Plan",
        "indent_plan_fsn_col": "Trichy FSN",
    },
    "chennai": {
        "label":           "Chennai",
        "tab_prefix":      "che",
        "sheet_id":        "1BquGJJri6WpJsUIre7JZLlpOeCR-etR3hLHYbHeOBOo", #main sheet link updated
        "indent_plan_tab": "Chennai Indent Plan",
        "indent_plan_fsn_col": "Chennai FSN",   # column name for FSN in the Indent Plan tab
    },
    "coimbatore": {
        "label":           "Coimbatore",
        "tab_prefix":      "coi",
        "sheet_id":        "19YLdB0JeEnTWEnvmVFVZIlc4D7T0eB1jRXotvSY6jJ8",  #main sheet link updated
        "indent_plan_tab": "Coimbatore Indent Plan",
        "indent_plan_fsn_col": "Coimbatore FSN",
    },
    "bengaluru": {
        "label":           "Bengaluru",
        "tab_prefix":      "ben",
        "sheet_id":        "1F2S90yz-rHIDCfAAVXi4p0tQBuDO5fRlm-5T8lzlMFw",
        "indent_plan_sheet_id": "1LR-UGBA9iOdrQ5eMm1ndHwV_w3rhVwtAcPBJJ__69dA",
        "indent_plan_tab": "Bangalore Indent Plan",
        "indent_plan_fsn_col": "Bangalore FSN",
    },
    "mumbai": {
        "label":           "Mumbai",
        "tab_prefix":      "mum",
        "sheet_id":        "1y2LaBblsqRLX1lG0XpOTVzUGvE40q_z_ZIcLyxPpl54",
        "indent_plan_sheet_id": "1e3rd1kClSqWMg7ewfG-aPAUNyY4puzlGvfh126gxgEA", #main indent sheet ID
        "indent_plan_tab": "Mumbai Indent Plan",
        "indent_plan_fsn_col": "Mumbai FSN",
    },
}

if os.environ.get("VERCEL"):
    OUTPUT_DIR = "/tmp"
else:
    OUTPUT_DIR = "outputs"
    os.makedirs(OUTPUT_DIR, exist_ok=True)

# ─────────────────────────────────────────────
# COLOURS
# ─────────────────────────────────────────────
CLR = {
    "header_bg":   "FFF2CC",   # light yellow to match user image
    "header_fg":   "000000",   # black text
    "match_bg":    "C6EFCE",   # green  ✅
    "match_fg":    "276221",
    "mismatch_bg": "FFC7CE",   # red    ❌
    "mismatch_fg": "9C0006",
    "section_bg":  "BDD7EE",   # light blue – pivot headers
    "alt_row":     "FFFFFF",   # white background for rows
    "subtotal_bg": "DDEBF7",   # slightly darker blue for subtotals
    "total_bg":    "FFE699",   # yellow – grand total
}

thin = Side(style="medium", color="000000")  # A medium black border matches the screenshot better
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)


# ─────────────────────────────────────────────
# GOOGLE SHEETS FETCH
# ─────────────────────────────────────────────
def fetch_indent_sheet(target_date: date, city_key: str) -> pd.DataFrame:
    """
    Fetches the tab  {prefix}_YYYY-MM-DD_PO  from the city's Indent Google Sheet.
    Returns a DataFrame with columns:
        FSN, Qty, SLA, Warehouse, Brand, Store, FSN_Title, VERTICAL

    city_key : one of 'trichy', 'chennai', 'coimbatore'
    """
    city_cfg = CITIES[city_key]
    tab_name = f"{city_cfg['tab_prefix']}_{target_date.strftime('%Y-%m-%d')}_PO"
    sheet_id = city_cfg["sheet_id"]

    if "YOUR_" in sheet_id:
        raise RuntimeError(
            f"Sheet ID for '{city_cfg['label']}' is not configured yet. "
            f"Please update CITIES['{city_key}']['sheet_id'] in the script."
        )

    try:
        import gspread
        from google.oauth2.service_account import Credentials

        scopes = [
            "https://www.googleapis.com/auth/spreadsheets.readonly",
            "https://www.googleapis.com/auth/drive.readonly",
        ]
        creds = Credentials.from_service_account_file(
            SERVICE_ACCOUNT_KEY, scopes=scopes
        )
        gc = gspread.authorize(creds)
        sh = gc.open_by_key(sheet_id)
        ws = sh.worksheet(tab_name)
        all_values = ws.get_all_values()
        if all_values:
            headers = all_values[0]
            # De-duplicate / fill empty headers so DataFrame doesn't complain
            seen = {}
            clean_headers = []
            for i, h in enumerate(headers):
                h = h.strip()
                if h == "":
                    h = f"col_{i}"
                if h in seen:
                    seen[h] += 1
                    h = f"{h}_{seen[h]}"
                else:
                    seen[h] = 0
                clean_headers.append(h)
            df = pd.DataFrame(all_values[1:], columns=clean_headers)
            for c in df.columns:
                df[c] = df[c].astype(str).str.strip()
        else:
            df = pd.DataFrame()

    except ImportError:
        # ── OFFLINE FALLBACK (for testing without gspread installed) ──
        print(f"[WARN] gspread not available. Using offline sample data for {tab_name}.")
        df = _sample_indent_data(target_date)

    except RuntimeError:
        raise

    except Exception as e:
        raise RuntimeError(f"Could not fetch Google Sheet tab '{tab_name}': {e}")

    df.columns = [c.strip() for c in df.columns]
    # Normalise column names to match XD allocation sheet
    rename_map = {
        "FSN Title":   "FSN_Title",
        "FSN Title ":  "FSN_Title",
        # Chennai column names
        "Title":       "FSN_Title",
        "Vertical":    "VERTICAL",
        "Store ID":    "Store",
        "Supplier ID": "Supplier ID",
        "Store Site ID": "Warehouse",
        "Title ":      "FSN_Title",
        # Quantity column variations across cities
        "PO qty":      "Qty",
        "PO Qty":      "Qty",
        "PO QTY":      "Qty",
        "Quantity":    "Qty",
        "quantity":    "Qty",
        "qty":         "Qty",
        "QTY":         "Qty",
    }
    df = df.rename(columns=rename_map)
    df["Qty"] = pd.to_numeric(df["Qty"], errors="coerce").fillna(0)
    return df

def fetch_indent_plan(target_date: date, city_key: str):
    """
    Fetches the city's Indent Plan tab (e.g. 'Chennai Indent Plan'),
    filters rows where  Po Date (D-1)  matches  target_date,
    and returns a DataFrame with columns:
        FSN, Indent_Qty, VERTICAL, Brand, FSN_Title, plus date columns.

    Returns None if the city has no indent_plan_tab configured.
    """
    city_cfg = CITIES[city_key]
    plan_tab = city_cfg.get("indent_plan_tab")
    if not plan_tab:
        return None

    sheet_id    = city_cfg.get("indent_plan_sheet_id") or city_cfg["sheet_id"]
    fsn_col_raw = city_cfg.get("indent_plan_fsn_col", "FSN")

    if "YOUR_" in sheet_id:
        print(f"[WARN] Indent Plan Sheet ID for '{city_cfg['label']}' is not configured yet.")
        return None

    try:
        import gspread
        from google.oauth2.service_account import Credentials

        scopes = [
            "https://www.googleapis.com/auth/spreadsheets.readonly",
            "https://www.googleapis.com/auth/drive.readonly",
        ]
        creds = Credentials.from_service_account_file(SERVICE_ACCOUNT_KEY, scopes=scopes)
        gc    = gspread.authorize(creds)
        sh    = gc.open_by_key(sheet_id)
        ws    = sh.worksheet(plan_tab)
        all_values = ws.get_all_values()
        if not all_values:
            return None

        # Build DataFrame
        headers = [h.strip() for h in all_values[0]]
        df = pd.DataFrame(all_values[1:], columns=headers)
        for c in df.columns:
            df[c] = df[c].astype(str).str.strip()

        # ── Normalise FSN column ──────────────────────────────────────────
        if fsn_col_raw in df.columns:
            df = df.rename(columns={fsn_col_raw: "FSN"})
        elif "FSN" not in df.columns:
            raise RuntimeError(f"Could not find FSN column '{fsn_col_raw}' in '{plan_tab}'")

        # ── Normalise other column names ──────────────────────────────────
        col_rename = {
            "PO qty":            "Indent_Qty",
            "PO Qty":            "Indent_Qty",
            "PO QTY":            "Indent_Qty",
            "Vertical":          "VERTICAL",
            "Title":             "FSN_Title",
            "FSN Title":         "FSN_Title",
            "Po Date (D-1)":     "PO_Date",
            "Po date (D-1)":     "PO_Date",
            "PO Date (D-1)":     "PO_Date",
        }
        df = df.rename(columns=col_rename)

        # ── Parse & filter by PO date ─────────────────────────────────────
        # Dates in the sheet are stored as MM/DD/YYYY strings
        df["PO_Date_Parsed"] = pd.to_datetime(df["PO_Date"], dayfirst=False, errors="coerce").dt.date
        df = df[df["PO_Date_Parsed"] == target_date].copy()

        if df.empty:
            print(f"[WARN] No rows found in '{plan_tab}' for PO date {target_date}")
            return None

        df["Indent_Qty"] = pd.to_numeric(df["Indent_Qty"], errors="coerce").fillna(0)

        # ── Identify and parse date columns to retain ─────────────────────
        date_cols = {}
        for c in df.columns:
            cl = c.strip().lower()
            if "indent_date" in cl or "indent date" in cl:
                date_cols["Indent_date (D-2)"] = c
            elif "po date" in cl or "po_date" in cl:
                date_cols["Po Date (D-1)"] = c
            elif "ds delivery" in cl or "delivery date" in cl:
                date_cols["DS Delivery Date"] = c
            elif "store live" in cl or "live date" in cl:
                date_cols["Store Live Date"] = c

        for std_col, raw_col in date_cols.items():
            df[std_col] = pd.to_datetime(df[raw_col], errors="coerce").dt.date

        # ── Group by FSN ──────────────────────────────────────────────────
        agg_dict = {
            "Indent_Qty": "sum",
            "VERTICAL": "first",
            "Brand": "first",
            "FSN_Title": "first",
        }
        for std_col in date_cols.keys():
            agg_dict[std_col] = "first"

        grp = df.groupby("FSN", as_index=False).agg(agg_dict)
        return grp

    except RuntimeError:
        raise
    except Exception as e:
        print(f"[WARN] Could not fetch Indent Plan for {city_cfg['label']}: {e}")
        return None

def _sample_indent_data(target_date: date) -> pd.DataFrame:
    """Minimal offline sample – mirrors the structure of Trichy_Indent_sheet.xlsx."""
    rows = [
        ("BTMHJ4JZYXCBJYZA", 61,  "Buttermilk & Lassi", "Rta_106_Bhima Nagar",    "Hatsun",     "rta_106_wh_hl_01", "Hatsun buttermilk 200ML",           "Buttermilk & Lassi"),
        ("CUYHBFBHAUR6YM5X", 6,   "CurdYogurt",         "Rta_106_Bhima Nagar",    "Hatsun",     "rta_106_wh_hl_01", "Hatsun Cup Curd 200 g",             "CurdYogurt"),
        ("CUYHBFBHEGGYHC3W", 120, "CurdYogurt",         "Rta_109_Ashok Nagar",    "Hatsun",     "rta_109_wh_hl_01", "Hatsun Pouch Curd 500 g",           "CurdYogurt"),
        ("MLKHBF78BMFPHABN", 56,  "Milk",               "Rta_106_Bhima Nagar",    "Arokya",     "rta_106_wh_hl_01", "Arokya Full Cream Milk 500 ml",     "Milk"),
        ("MLKHBF78YDFVJAC7", 84,  "Milk",               "Rta_106_Bhima Nagar",    "Arokya",     "rta_106_wh_hl_01", "Arokya Standardised Milk 500 ml",   "Milk"),
        ("MLKHBF78KZYQXWJN", 56,  "Milk",               "Rta_106_Bhima Nagar",    "Arokya",     "rta_106_wh_hl_01", "Arokya Toned Milk 500 ml",          "Milk"),
        ("PTFHBFBX4YUXZR8S", 72,  "PaneerTofu",         "Rta_113_Thanjavur",      "Hatsun",     "rta_113_wh_hl_01", "Hatsun Paneer 200 g",               "PaneerTofu"),
        ("RMDHVHYVJPBBDWY",  30,  "ReadyMeals",         "Rta_106_Bhima Nagar",    "Elanadu",    "rta_106_wh_hl_01", "ELANADU HALF COOKED CHAPPATHI 400GM","ReadyMeals"),
        ("RMDHVHYNYQUK8Z9",  30,  "ReadyMeals",         "Rta_106_Bhima Nagar",    "Elanadu",    "rta_106_wh_hl_01", "Elanadu MALABAR POROTTA 400g",      "ReadyMeals"),
        ("RYMHHVKWTAEGZRBS", 20,  "ReadyMixes",         "Rta_106_Bhima Nagar",    "Elanadu",    "rta_106_wh_hl_01", "Elanadu DOSA IDLI BATTER 1kg",      "ReadyMixes"),
        ("BABZ38E47UEPAHJ",  4,   "Breads",             "Rta_109_Ashok Nagar",    "Fores Iniyaa","rta_109_wh_hl_01","Fores Iniyaa Wheat Bread 350g",     "Breads"),
        ("BABZ38EAUAKAHMR",  4,   "Breads",             "Rta_111_Gandhi Nagar",   "Fores Iniyaa","rta_111_wh_hl_01","Fores Iniyaa Jumbo Sandwich Bread 620g","Breads"),
        ("BABZ38EJ2G6XR4S",  6,   "Breads",             "Rta_113_Thanjavur",      "Fores Iniyaa","rta_113_wh_hl_01","Fores Iniyaa Sandwich Bread 400g",  "Breads"),
        ("BABZ38EU38NPHRJ",  5,   "Breads",             "Rta_107_Kajamalai Colony","Fores Iniyaa","rta_107_wh_hl_01","Fores Iniyaa Milk Bread 300g",     "Breads"),
        ("BABZ38EUWHFZFGT",  4,   "Breads",             "Rta_111_Gandhi Nagar",   "Fores Iniyaa","rta_111_wh_hl_01","Fores Iniyaa Sweet Bun 200g",       "Breads"),
    ]
    return pd.DataFrame(rows, columns=["FSN","Qty","SLA","Warehouse","Brand","Store","FSN_Title","VERTICAL"])


# ─────────────────────────────────────────────
# DATA PROCESSING
# ─────────────────────────────────────────────
def build_po_sheet(indent_df: pd.DataFrame, target_date: date = None, city_key: str = "") -> pd.DataFrame:
    """
    The PO sheet = raw store-level allocation data pulled straight from
    the Indent Google Sheet (the same as fetch_indent_sheet output).
    """
    po = indent_df.copy()
    return po


def build_indent_summary(indent_df: pd.DataFrame, plan_df=None, target_date: date = None) -> pd.DataFrame:
    """
    Builds the Indent Summary sheet:
      - One row per FSN (all verticals included)
      - Based on the Indent Plan (plan_df) if available, showing all planned items.
      - PO_Qty     = SUM of Qty from the daily PO tab (per FSN)
      - Indent_Qty = SUM from the Indent Plan tab filtered by PO date
                     (falls back to PO_Qty if Indent Plan not available)
      - Difference = Indent_Qty - PO_Qty
      - Match      = checkmark or cross
    """
    date_cols = ["Indent_date (D-2)", "Po Date (D-1)", "DS Delivery Date", "Store Live Date"]

    if plan_df is not None and not plan_df.empty:
        # Get PO totals from daily PO tab per FSN, excluding eggs
        po_non_eggs = indent_df[~indent_df["VERTICAL"].str.strip().str.lower().isin(["eggs", "egg"])].copy()
        po_totals = (
            po_non_eggs.groupby("FSN", as_index=False)
            .agg(
                PO_Qty=("Qty", "sum"),
                Brand=("Brand", "first"),
                FSN_Title=("FSN_Title", "first"),
                VERTICAL=("VERTICAL", "first"),
            )
        )
        
        # Determine plan columns to merge
        merge_cols = ["FSN", "Indent_Qty", "Brand", "FSN_Title", "VERTICAL"]
        for c in date_cols:
            if c in plan_df.columns:
                merge_cols.append(c)

        # Merge using outer join to capture FSNs present in either sheet
        summary = po_totals.merge(
            plan_df[merge_cols],
            on="FSN",
            how="outer",
            suffixes=("", "_plan")
        )
        
        # Coalesce / fill NA
        summary["PO_Qty"] = summary["PO_Qty"].fillna(0)
        summary["Indent_Qty"] = summary["Indent_Qty"].fillna(0)
        summary["Brand"] = summary["Brand"].fillna(summary["Brand_plan"])
        summary["FSN_Title"] = summary["FSN_Title"].fillna(summary["FSN_Title_plan"])
        summary["VERTICAL"] = summary["VERTICAL"].fillna(summary["VERTICAL_plan"])
        summary = summary.drop(columns=["Brand_plan", "FSN_Title_plan", "VERTICAL_plan"])

        # Coalesce dates
        if target_date is not None:
            if "Indent_date (D-2)" in summary.columns:
                summary["Indent_date (D-2)"] = summary["Indent_date (D-2)"].fillna(target_date - timedelta(days=1))
            if "Po Date (D-1)" in summary.columns:
                summary["Po Date (D-1)"] = summary["Po Date (D-1)"].fillna(target_date)
            if "DS Delivery Date" in summary.columns:
                summary["DS Delivery Date"] = summary["DS Delivery Date"].fillna(target_date + timedelta(days=1))
            if "Store Live Date" in summary.columns:
                summary["Store Live Date"] = summary["Store Live Date"].fillna(target_date + timedelta(days=1))
    else:
        # Fallback: base on PO tab, excluding eggs
        po_non_eggs = indent_df[~indent_df["VERTICAL"].str.strip().str.lower().isin(["eggs", "egg"])].copy()
        summary = (
            po_non_eggs.groupby("FSN", as_index=False)
            .agg(
                PO_Qty=("Qty",      "sum"),
                Brand=("Brand",      "first"),
                FSN_Title=("FSN_Title", "first"),
                VERTICAL=("VERTICAL",  "first"),
            )
        )
        summary["Indent_Qty"] = summary["PO_Qty"]
        if target_date is not None:
            summary["Indent_date (D-2)"] = target_date - timedelta(days=1)
            summary["Po Date (D-1)"] = target_date
            summary["DS Delivery Date"] = target_date + timedelta(days=1)
            summary["Store Live Date"] = target_date + timedelta(days=1)

    summary["Difference"] = summary["Indent_Qty"] - summary["PO_Qty"]
    summary["Match"] = summary["Difference"].apply(
        lambda x: "\u2705 Match" if x == 0 else "\u274c Mismatch"
    )

    # Determine column ordering
    cols = [c for c in date_cols if c in summary.columns]
    cols += ["FSN", "Brand", "FSN_Title", "VERTICAL", "PO_Qty", "Indent_Qty", "Difference", "Match"]
    return summary[cols].sort_values(["VERTICAL", "FSN"]).reset_index(drop=True)


def build_pivot_vertical(indent_df: pd.DataFrame, city_key: str = "trichy") -> pd.DataFrame:
    if city_key in ["bengaluru", "mumbai"]:
        # Group by Supplier ID (SLA) and VERTICAL
        supp_col = "Supplier ID" if "Supplier ID" in indent_df.columns else "SLA"
        
        # Calculate sums per group
        grouped = indent_df.groupby([supp_col, "VERTICAL"], as_index=False)["Qty"].sum()
        
        # Calculate sub-totals per Supplier ID
        sub_totals = indent_df.groupby(supp_col, as_index=False)["Qty"].sum()
        
        # Build the final hierarchical list
        rows = []
        for _, st in sub_totals.iterrows():
            s_id = st[supp_col]
            s_qty = st["Qty"]
            # Add Supplier ID header row
            rows.append([s_id, s_qty, ""])
            # Add Vertical items under this supplier
            s_items = grouped[grouped[supp_col] == s_id]
            for _, item in s_items.iterrows():
                rows.append([f"  {item['VERTICAL']}", item["Qty"], ""])
                
        # Create DataFrame
        pv = pd.DataFrame(rows, columns=["Row Labels", "Sum of QTY", "Mismatch Error"])
        grand_total = pd.DataFrame([["Grand Total", sub_totals["Qty"].sum(), ""]], columns=pv.columns)
        return pd.concat([pv, grand_total], ignore_index=True)
    else:
        pv = indent_df.groupby("VERTICAL", as_index=False)["Qty"].sum()
        pv.columns = ["Vertical", "Total_Qty"]
        pv = pv.sort_values("Total_Qty", ascending=False).reset_index(drop=True)
        grand = pd.DataFrame([["Grand Total", pv["Total_Qty"].sum()]], columns=pv.columns)
        return pd.concat([pv, grand], ignore_index=True)


def build_pivot_fsn(indent_df: pd.DataFrame) -> pd.DataFrame:
    pv = indent_df.groupby(["FSN", "FSN_Title", "VERTICAL"], as_index=False)["Qty"].sum()
    pv.columns = ["FSN", "FSN_Title", "VERTICAL", "Total_Qty"]
    pv = pv.sort_values(["VERTICAL", "Total_Qty"], ascending=[True, False]).reset_index(drop=True)
    grand = pd.DataFrame([["", "Grand Total", "", pv["Total_Qty"].sum()]], columns=pv.columns)
    return pd.concat([pv, grand], ignore_index=True)


def build_mismatch_report(indent_summary: pd.DataFrame) -> pd.DataFrame:
    mismatches = indent_summary[indent_summary["Match"] == "❌ Mismatch"].copy()
    return mismatches.reset_index(drop=True)


def build_store_summary(indent_df: pd.DataFrame) -> tuple:
    """
    Returns:
      store_df   – distinct stores with their warehouse code, verticals served,
                   FSN count, and total qty allocated to that store
      total_count – int, number of unique stores
    """
    # Distinct store list with enriched info
    store_df = (
        indent_df.groupby("Store", as_index=False)
        .agg(
            Warehouse=("Warehouse", "first"),
            Verticals_Served=("VERTICAL", lambda x: ", ".join(sorted(x.unique()))),
            FSN_Count=("FSN", "nunique"),
            Total_Qty=("Qty", "sum"),
        )
        .sort_values("Store")
        .reset_index(drop=True)
    )
    store_df.columns = ["Store", "Warehouse", "Verticals Served", "FSN Count", "Total Qty"]
    total_count = len(store_df)
    return store_df, total_count


def build_egg_pivot(po_df: pd.DataFrame) -> pd.DataFrame:
    eggs = po_df[po_df["VERTICAL"].str.lower() == "egg"].copy()
    if eggs.empty:
        return pd.DataFrame(columns=["Store Site ID", "Title", "Sum of QTY"])
    
    eggs["Qty"] = pd.to_numeric(eggs["Qty"], errors="coerce").fillna(0)
    grouped = eggs.groupby(["Warehouse", "FSN_Title"], as_index=False)["Qty"].sum()
    
    rows = []
    for store in grouped["Warehouse"].unique():
        store_data = grouped[grouped["Warehouse"] == store]
        for _, row in store_data.iterrows():
            rows.append([row["Warehouse"], row["FSN_Title"], row["Qty"]])
        
        total_qty = store_data["Qty"].sum()
        rows.append(["Total", "", total_qty])
        
    grand_total = eggs["Qty"].sum()
    rows.append(["Grand Total", "", grand_total])
        
    return pd.DataFrame(rows, columns=["Store Site ID", "Title", "Sum of QTY"])


# ─────────────────────────────────────────────
# EXCEL WRITING HELPERS
# ─────────────────────────────────────────────
def _hdr_style(cell, bg=CLR["header_bg"], fg=CLR["header_fg"]):
    cell.font      = Font(bold=True, color=fg, name="Arial", size=10)
    cell.fill      = PatternFill("solid", start_color=bg)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border    = BORDER


def _data_style(cell, bold=False, bg=None, fg="000000", num_fmt=None, align="left"):
    cell.font      = Font(bold=bold, color=fg, name="Arial", size=10)
    cell.alignment = Alignment(horizontal=align, vertical="center")
    cell.border    = BORDER
    if bg:
        cell.fill = PatternFill("solid", start_color=bg)
    if num_fmt:
        cell.number_format = num_fmt


def _auto_width(ws, min_w=10, max_w=40):
    for col in ws.columns:
        col_letter = get_column_letter(col[0].column)
        length = max((len(str(c.value or "")) for c in col), default=min_w)
        ws.column_dimensions[col_letter].width = min(max(length + 2, min_w), max_w)


def _write_df_to_sheet(ws, df: pd.DataFrame, title_row: str = None,
                        match_col: str = None, total_last_row: bool = False,
                        formula_map: dict = None):
    """Generic helper – writes header + data rows with formatting."""
    if formula_map is None:
        formula_map = {}

    start_row = 1
    if title_row:
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(df.columns))
        tc = ws.cell(row=1, column=1, value=title_row)
        tc.font  = Font(bold=True, color=CLR["header_fg"], name="Arial", size=12)
        tc.fill  = PatternFill("solid", start_color=CLR["header_bg"])
        tc.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 22
        start_row = 2

    # Header
    for col_idx, col_name in enumerate(df.columns, 1):
        _hdr_style(ws.cell(row=start_row, column=col_idx, value=col_name))
    ws.row_dimensions[start_row].height = 20

    # Data rows
    for row_idx, row in enumerate(df.itertuples(index=False), start=start_row + 1):
        is_total = total_last_row and row_idx == start_row + len(df)
        is_subtotal = (str(row[0]).endswith(" Total") or str(row[0]) == "Total") and not is_total
        is_alt   = (row_idx % 2 == 0)

        # Precompute formula formatting arguments for dynamic cells
        fmt_args = {"row": row_idx}
        for c_idx, c_name in enumerate(df.columns, 1):
            fmt_args[f"{c_name}_col"] = get_column_letter(c_idx)
            fmt_args[f"{c_name}_cell"] = f"{get_column_letter(c_idx)}{row_idx}"

        for col_idx, value in enumerate(row, 1):
            col_name = df.columns[col_idx - 1]

            # Inject formula if mapped
            if not is_total and not is_subtotal and col_name in formula_map:
                try:
                    value = formula_map[col_name].format(**fmt_args)
                except KeyError:
                    pass

            cell = ws.cell(row=row_idx, column=col_idx, value=value)

            # colour match / mismatch column
            if match_col and col_name == match_col:
                if str(value).startswith("✅"):
                    _data_style(cell, bg=CLR["match_bg"], fg=CLR["match_fg"], bold=True, align="center")
                elif str(value).startswith("❌"):
                    _data_style(cell, bg=CLR["mismatch_bg"], fg=CLR["mismatch_fg"], bold=True, align="center")
                else:
                    _data_style(cell, align="center")
                continue

            import datetime as dt
            is_date_col = "date" in col_name.lower()

            if is_total:
                _data_style(cell, bold=True, bg=CLR["total_bg"], align="right" if isinstance(value, (int, float)) or str(value).startswith("=") else "left")
            elif is_subtotal:
                _data_style(cell, bold=True, bg=CLR["subtotal_bg"], align="right" if isinstance(value, (int, float)) or str(value).startswith("=") else "left")
            elif isinstance(value, (int, float)) or str(value).startswith("="):
                _data_style(cell, bg=None if not is_alt else CLR["alt_row"], align="right", num_fmt="#,##0")
            elif isinstance(value, (dt.date, dt.datetime)) or is_date_col:
                _data_style(cell, bg=None if not is_alt else CLR["alt_row"], align="center", num_fmt="mm-dd-yy")
            else:
                _data_style(cell, bg=None if not is_alt else CLR["alt_row"])


# ─────────────────────────────────────────────
# MAIN REPORT BUILDER
# ─────────────────────────────────────────────
def build_excel_report(target_date: date, output_path: str, city_key: str = "trichy") -> str:
    city_label = CITIES[city_key]["label"]
    print(f"[{city_label}] [1/7] Fetching Indent sheet for {target_date} ...")
    indent_df = fetch_indent_sheet(target_date, city_key)

    print(f"[{city_label}] [2/7] Building PO sheet ...")
    po_df = build_po_sheet(indent_df, target_date, city_key)

    print(f"[{city_label}] [3/7] Building Indent summary (VLOOKUP + cross-verify) ...")
    plan_df = fetch_indent_plan(target_date, city_key)
    indent_summary = build_indent_summary(indent_df, plan_df, target_date)

    print(f"[{city_label}] [4/7] Building pivot tables ...")
    pv_vertical = build_pivot_vertical(indent_df, city_key)
    pv_fsn      = build_pivot_fsn(indent_df)

    print(f"[{city_label}] [5/7] Building mismatch report ...")
    mismatches  = build_mismatch_report(indent_summary)

    print(f"[{city_label}] [6/7] Building store summary ...")
    store_df, total_stores = build_store_summary(indent_df)

    # Store comparison vs Yesterday
    try:
        from datetime import timedelta
        yday = target_date - timedelta(days=1)
        yday_df = fetch_indent_sheet(yday, city_key)
        yday_stores = set(yday_df["Store"].dropna().unique()) if "Store" in yday_df.columns else set()
        today_stores = set(indent_df["Store"].dropna().unique()) if "Store" in indent_df.columns else set()
        added_stores = sorted(list(today_stores - yday_stores))
        removed_stores = sorted(list(yday_stores - today_stores))
        store_compare_status = "ok"
    except Exception as e:
        added_stores = []
        removed_stores = []
        store_compare_status = str(e)

    print(f"[{city_label}] [7/7] Writing Excel ...")
    wb = Workbook()
    wb.remove(wb.active)   # remove default sheet

    date_str = target_date.strftime("%d-%b-%Y")
    city_label = CITIES[city_key]["label"]

    # ── Sheet 1 : PO (store-level allocation) ──────────────────────────
    ws_po = wb.create_sheet("PO")
    
    # Format PO sheet according to city requirements
    po_out = po_df.copy()
    supp_col = "Supplier ID" if "Supplier ID" in po_out.columns else "SLA"
    
    if city_key in ["bengaluru", "mumbai"]:
        po_out["City"] = po_out.get("City", "Bengaluru" if city_key == "bengaluru" else "Mumbai")
        po_out["Store ID"] = po_out.get("Store", "")
        po_out["Store Site ID"] = po_out.get("Warehouse", po_out.get("Store Site ID", ""))
        po_out["QTY"] = po_out.get("Qty", "")
        po_out["SLA"] = po_out.get("SLA", po_out.get(supp_col, ""))
        po_out["Supplier ID"] = po_out.get("Supplier ID", po_out.get(supp_col, ""))
        po_out["Contract ID"] = po_out.get("Contract ID", "")
        po_out["Po No"] = po_out.get("Po No", "")
        po_out["Title"] = po_out.get("Title", po_out.get("FSN_Title", ""))
        po_out["Brand"] = po_out.get("Brand", "")
        
        # Ensure we properly fallback without NaNs overwriting
        vert_series = po_out.get("Vertical")
        if vert_series is None or vert_series.isna().all():
            vert_series = po_out.get("VERTICAL", "")
        po_out["Vertical"] = vert_series
        
        if city_key == "mumbai":
            tag_series = po_out.get("tag")
            if tag_series is None or tag_series.isna().all():
                tag_series = po_out.get("Chiller/Non chiller Tag", "")
            po_out["tag"] = tag_series
            out_cols = ["City", "Store ID", "Store Site ID", "FSN", "QTY", "SLA", "Supplier ID", "Contract ID", "Po No", "Title", "Brand", "Vertical", "tag"]
        else: # bengaluru
            tag_series = po_out.get("Chiller/Non chiller Tag")
            if tag_series is None or tag_series.isna().all():
                tag_series = po_out.get("tag", "")
            po_out["Chiller/Non chiller Tag"] = tag_series
            out_cols = ["City", "Store ID", "Store Site ID", "FSN", "QTY", "SLA", "Supplier ID", "Contract ID", "Po No", "Title", "Brand", "Vertical", "Chiller/Non chiller Tag"]
            
        for c in out_cols:
            if c not in po_out.columns:
                po_out[c] = ""
                
        po_out = po_out[out_cols].fillna("")
        po_out = po_out.sort_values(["Vertical", "FSN", "Store ID"]).reset_index(drop=True)
    else:
        po_out["Date"] = target_date.strftime("%d-%b-%Y") if target_date else ""
        po_out["Contract ID"] = po_out.get("Contract ID", "")
        po_out["Warehouse"] = po_out.get("Warehouse", "")
        
        po_out = po_out.rename(columns={
            "FSN_Title": "Title",
            "Qty": "PO qty",
            "VERTICAL": "Vertical",
            supp_col: "Supplier ID",
            "Store": "Store ID",
        })
        
        out_cols = ["Date", "Brand", "Title", "FSN", "PO qty", "Vertical", "Supplier ID", "Contract ID", "Store ID", "Warehouse"]
        for c in out_cols:
            if c not in po_out.columns:
                po_out[c] = ""
                
        po_out = po_out[out_cols]
        po_out = po_out.sort_values(["Vertical", "FSN", "Warehouse"]).reset_index(drop=True)
        
    _write_df_to_sheet(ws_po, po_out)
    _auto_width(ws_po)


    # ── Sheet 1b : vertical level difference ────────────────────────────
    ws_vld = wb.create_sheet("vertical level difference")
    # Headers on row 3
    _hdr_style(ws_vld.cell(row=3, column=1, value="Row Labels"))
    _hdr_style(ws_vld.cell(row=3, column=2, value="PO Qty"))
    _hdr_style(ws_vld.cell(row=3, column=3, value="Indent Qty"))
    _hdr_style(ws_vld.cell(row=3, column=4, value="Diff"))
    ws_vld.row_dimensions[3].height = 24

    # Unique verticals from PO, standardized (stripped, title cased, and deduplicated)
    all_verticals = sorted(list(set([str(v).strip().title() for v in indent_df["VERTICAL"].dropna() if str(v).strip() != ""])))
    
    # Find column letters dynamically based on header names
    # In PO sheet:
    po_out_cols = list(po_out.columns)
    qty_col_name = "QTY" if "QTY" in po_out_cols else ("PO qty" if "PO qty" in po_out_cols else "Qty")
    vert_col_name = "Vertical" if "Vertical" in po_out_cols else "VERTICAL"
    
    po_qty_col = get_column_letter(po_out_cols.index(qty_col_name) + 1)
    po_vert_col = get_column_letter(po_out_cols.index(vert_col_name) + 1)

    # In Indent Summary sheet:
    is_cols = list(indent_summary.columns)
    is_qty_col = get_column_letter(is_cols.index("Indent_Qty") + 1)
    is_vert_col = get_column_letter(is_cols.index("VERTICAL") + 1)

    row_idx = 4
    for vert in all_verticals:
        is_alt = (row_idx % 2 == 0)
        # Row Labels (Vertical)
        c1 = ws_vld.cell(row=row_idx, column=1, value=vert)
        _data_style(c1, bg=CLR["alt_row"] if is_alt else None)
        
        # Sum of Qty (from PO sheet using SUMIFS)
        c2 = ws_vld.cell(row=row_idx, column=2, value=f"=SUMIFS(PO!{po_qty_col}:{po_qty_col},PO!{po_vert_col}:{po_vert_col},A{row_idx})")
        _data_style(c2, bg=CLR["alt_row"] if is_alt else None, align="right", num_fmt="#,##0")
        
        # Total PO Qty (Excluding Eggs) (from Indent Summary using SUMIFS)
        c3 = ws_vld.cell(row=row_idx, column=3, value=f"=SUMIFS('Indent Summary'!{is_qty_col}:{is_qty_col},'Indent Summary'!{is_vert_col}:{is_vert_col},A{row_idx})")
        _data_style(c3, bg=CLR["alt_row"] if is_alt else None, align="right", num_fmt="#,##0")
        
        # Diff (Sum of Qty - Total PO Qty)
        c4 = ws_vld.cell(row=row_idx, column=4, value=f"=B{row_idx}-C{row_idx}")
        _data_style(c4, bg=CLR["alt_row"] if is_alt else None, align="right", num_fmt="#,##0")
        
        row_idx += 1

    # Grand Total row
    ws_vld.cell(row=row_idx, column=1, value="Grand Total")
    ws_vld.cell(row=row_idx, column=2, value=f"=SUM(B4:B{row_idx-1})")
    ws_vld.cell(row=row_idx, column=3, value=f"=SUM(C4:C{row_idx-1})")
    ws_vld.cell(row=row_idx, column=4, value=f"=B{row_idx}-C{row_idx}")
    for col_idx in range(1, 5):
        cell = ws_vld.cell(row=row_idx, column=col_idx)
        _data_style(cell, bold=True, bg=CLR["total_bg"],
                    align="right" if col_idx > 1 else "left",
                    num_fmt="#,##0" if col_idx > 1 else None)
    
    _auto_width(ws_vld)

    # ── Sheet 2 : Indent Summary ────────────────────────────────────────
    ws_indent = wb.create_sheet("Indent Summary")
    
    po_out_cols = list(po_out.columns)
    qty_col_name = "QTY" if "QTY" in po_out_cols else ("PO qty" if "PO qty" in po_out_cols else "Qty")
    fsn_col_name = "FSN"
    
    po_qty_col = get_column_letter(po_out_cols.index(qty_col_name) + 1)
    po_fsn_col = get_column_letter(po_out_cols.index(fsn_col_name) + 1)
    
    indent_formula_map = {
        "PO_Qty": f"=SUMIFS('PO'!${po_qty_col}:${po_qty_col}, 'PO'!${po_fsn_col}:${po_fsn_col}, {{FSN_cell}})",
        "Difference": "={Indent_Qty_cell}-{PO_Qty_cell}",
        "Match": '=IF({Difference_cell}=0, "✅ Match", "❌ Mismatch")'
    }

    _write_df_to_sheet(
        ws_indent, indent_summary,
        # title_row=f"Indent Cross-Verification  |  {date_str}",
        match_col="Match",
        total_last_row=False,
        formula_map=indent_formula_map
    )
    _auto_width(ws_indent)

    # Add a totals row at the bottom of Indent Summary
    # No title row (it's commented out), so: row 1 = header, row 2 onwards = data
    last_data_row = len(indent_summary) + 1   # +1 header only
    total_row     = last_data_row + 1
    ws_indent.cell(row=total_row, column=1, value="GRAND TOTAL")
    for col_idx, col_name in enumerate(indent_summary.columns, 1):
        cell = ws_indent.cell(row=total_row, column=col_idx)
        if col_name in ("PO_Qty", "Indent_Qty", "Difference"):
            col_letter = get_column_letter(col_idx)
            cell.value = f"=SUM({col_letter}2:{col_letter}{last_data_row})"
        _data_style(cell, bold=True, bg=CLR["total_bg"],
                    align="right" if col_name in ("PO_Qty","Indent_Qty","Difference") else "left")

    # ── Sheet 3 : Pivot by Vertical ─────────────────────────────────────
    ws_pv1 = wb.create_sheet("Pivot – Vertical")
    
    po_vert_col = get_column_letter(po_out_cols.index("Vertical" if "Vertical" in po_out_cols else "VERTICAL") + 1)
    
    formula_map_pv1 = {}
    if city_key not in ["mumbai", "bengaluru"]:
        formula_map_pv1["Total_Qty"] = f"=SUMIFS('PO'!${po_qty_col}:${po_qty_col}, 'PO'!${po_vert_col}:${po_vert_col}, {{Vertical_cell}})"
        
    _write_df_to_sheet(
        ws_pv1, pv_vertical,
        title_row=f"Qty by Vertical  |  {date_str}",
        total_last_row=True,
        formula_map=formula_map_pv1
    )
    
    if city_key in ["mumbai", "bengaluru"]:
        po_supp_col = "Supplier ID" if "Supplier ID" in po_out_cols else "SLA"
        po_supp_col_ltr = get_column_letter(po_out_cols.index(po_supp_col) + 1)
        
        current_supplier_cell = None
        last_row = 3 + len(pv_vertical) - 1 # excluding Grand Total
        for r in range(3, last_row):
            lbl = ws_pv1.cell(row=r, column=1).value
            if lbl and str(lbl).startswith("  "):
                ws_pv1.cell(row=r, column=2).value = f"=SUMIFS('PO'!${po_qty_col}:${po_qty_col}, 'PO'!${po_supp_col_ltr}:${po_supp_col_ltr}, {current_supplier_cell}, 'PO'!${po_vert_col}:${po_vert_col}, TRIM(A{r}))"
                ws_pv1.cell(row=r, column=3).value = f'=IF(COUNTIF($A$3:$A${last_row-1}, A{r})>1, "❌ Multiple Suppliers", "")'
                _data_style(ws_pv1.cell(row=r, column=3), align="center")
            else:
                current_supplier_cell = f"A{r}"
                ws_pv1.cell(row=r, column=2).value = f"=SUMIFS('PO'!${po_qty_col}:${po_qty_col}, 'PO'!${po_supp_col_ltr}:${po_supp_col_ltr}, A{r})"
                ws_pv1.cell(row=r, column=3).value = ""

    _auto_width(ws_pv1)

    # ── Sheet 4 : Pivot by FSN ──────────────────────────────────────────
    ws_pv2 = wb.create_sheet("Pivot – FSN")
    formula_map_pv2 = {
        "Total_Qty": f"=SUMIFS('PO'!${po_qty_col}:${po_qty_col}, 'PO'!${po_fsn_col}:${po_fsn_col}, {{FSN_cell}})"
    }
    _write_df_to_sheet(
        ws_pv2, pv_fsn,
        title_row=f"Qty by FSN  |  {date_str}",
        total_last_row=True,
        formula_map=formula_map_pv2
    )
    _auto_width(ws_pv2)

    # ── Sheet 5 : Store Summary ──────────────────────────────────────────
    ws_stores = wb.create_sheet("🏪 Store Summary")

    # Banner showing total distinct store count
    ws_stores.merge_cells("A1:E1")
    banner = ws_stores["A1"]
    banner.value     = f"Distinct Stores Active on {date_str}  —  Total: {total_stores} stores"
    banner.font      = Font(bold=True, color=CLR["header_fg"], name="Arial", size=12)
    banner.fill      = PatternFill("solid", start_color=CLR["header_bg"])
    banner.alignment = Alignment(horizontal="center", vertical="center")
    ws_stores.row_dimensions[1].height = 26

    # Header row at row 2
    for col_idx, col_name in enumerate(store_df.columns, 1):
        _hdr_style(ws_stores.cell(row=2, column=col_idx, value=col_name))
    ws_stores.row_dimensions[2].height = 20

    store_col_name = "Store ID" if "Store ID" in po_out_cols else "Store"
    if store_col_name not in po_out_cols:
        store_col_name = "Warehouse" # fallback
    po_store_col = get_column_letter(po_out_cols.index(store_col_name) + 1)

    # Data rows from row 3
    for row_idx, row in enumerate(store_df.itertuples(index=False), start=3):
        is_alt = (row_idx % 2 == 0)
        for col_idx, value in enumerate(row, 1):
            if store_df.columns[col_idx-1] == "Total Qty":
                value = f"=SUMIFS('PO'!${po_qty_col}:${po_qty_col}, 'PO'!${po_store_col}:${po_store_col}, A{row_idx})"
                
            cell = ws_stores.cell(row=row_idx, column=col_idx, value=value)
            if isinstance(value, (int, float)) or str(value).startswith("="):
                _data_style(cell, bg=CLR["alt_row"] if is_alt else None, align="right", num_fmt="#,##0")
            else:
                _data_style(cell, bg=CLR["alt_row"] if is_alt else None)

    # Grand total row
    total_row_s = 3 + len(store_df)
    totals = {"Store": "TOTAL", "Warehouse": "", "Verticals Served": "",
              "FSN Count": int(store_df["FSN Count"].sum()), "Total Qty": int(store_df["Total Qty"].sum())}
    for col_idx, col_name in enumerate(store_df.columns, 1):
        cell = ws_stores.cell(row=total_row_s, column=col_idx, value=totals[col_name])
        _data_style(cell, bold=True, bg=CLR["total_bg"],
                    align="right" if isinstance(totals[col_name], (int, float)) else "left",
                    num_fmt="#,##0" if isinstance(totals[col_name], (int, float)) else None)

    _auto_width(ws_stores)

    # ── Write Store Comparison ───────────────────────────────────────────
    ws_stores.cell(row=1, column=7, value="🟢 Added Since Yesterday").font = Font(bold=True, color="008000")
    ws_stores.cell(row=1, column=8, value="🔴 Removed Since Yesterday").font = Font(bold=True, color="FF0000")
    
    if store_compare_status == "ok":
        max_len = max(len(added_stores), len(removed_stores))
        if max_len == 0:
            ws_stores.cell(row=2, column=7, value="(No changes)").font = Font(italic=True, color="808080")
        else:
            for i in range(max_len):
                r_comp = i + 2
                if i < len(added_stores):
                    ws_stores.cell(row=r_comp, column=7, value=added_stores[i])
                if i < len(removed_stores):
                    ws_stores.cell(row=r_comp, column=8, value=removed_stores[i])
    else:
        ws_stores.cell(row=2, column=7, value="(Yesterday's data unavailable)").font = Font(italic=True, color="808080")
    
    ws_stores.column_dimensions[get_column_letter(7)].width = 25
    ws_stores.column_dimensions[get_column_letter(8)].width = 25

    if city_key in ["mumbai", "bengaluru"]:
        ws_eggs = wb.create_sheet("Eggs")
        egg_df = build_egg_pivot(po_df)
        
        wh_col_name = "Store Site ID" if "Store Site ID" in po_out_cols else "Warehouse"
        title_col_name = "Title" if "Title" in po_out_cols else "FSN_Title"
        
        po_wh_col = get_column_letter(po_out_cols.index(wh_col_name) + 1)
        po_title_col = get_column_letter(po_out_cols.index(title_col_name) + 1)
        
        formula_map_eggs = {
            "Sum of QTY": f"=SUMIFS('PO'!${po_qty_col}:${po_qty_col}, 'PO'!${po_wh_col}:${po_wh_col}, {{Store Site ID_cell}}, 'PO'!${po_title_col}:${po_title_col}, {{Title_cell}})"
        }
        
        _write_df_to_sheet(
            ws_eggs, egg_df,
            title_row=f"Eggs  |  {date_str}",
            total_last_row=True,
            formula_map=formula_map_eggs
        )
        _auto_width(ws_eggs)

    # ── Sheet 6 : Mismatch Report ────────────────────────────────────────
    ws_mm = wb.create_sheet("⚠ Mismatches")
    if mismatches.empty:
        ws_mm.merge_cells("A1:H1")
        c = ws_mm["A1"]
        c.value     = f"✅  All quantities matched for {date_str}. No mismatches found."
        c.font      = Font(bold=True, color=CLR["match_fg"], name="Arial", size=12)
        c.fill      = PatternFill("solid", start_color=CLR["match_bg"])
        c.alignment = Alignment(horizontal="center", vertical="center")
        ws_mm.row_dimensions[1].height = 30
    else:
        _write_df_to_sheet(
            ws_mm, mismatches,
            title_row=f"⚠ Mismatches Found  |  {date_str}",
            match_col="Match",
        )
        _auto_width(ws_mm)

    os.makedirs(os.path.dirname(output_path) if os.path.dirname(output_path) else ".", exist_ok=True)
    wb.save(output_path)
    print(f"[{city_label}] [OK] Report saved -> {output_path}")
    return output_path


# ─────────────────────────────────────────────
# DELEGATION WRAPPER
# ─────────────────────────────────────────────
def run_city_report(target_date: date, output_path: str, city_key: str) -> str:
    """
    Builds the report for all configured cities using the unified build_excel_report logic.
    """
    return build_excel_report(target_date, output_path, city_key)


# ─────────────────────────────────────────────
# EMBEDDED DASHBOARD HTML
# ─────────────────────────────────────────────
INDEX_HTML = """<!DOCTYPE html>
<html lang="en">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>XD Allocation Hub</title>
    <link href="https://fonts.googleapis.com/css2?family=Inter:wght@400;500;600;700&display=swap" rel="stylesheet">
    <style>
        *, *::before, *::after { box-sizing: border-box; margin: 0; padding: 0; }

        body {
            font-family: 'Inter', sans-serif;
            background: #f5f6f8;
            color: #111827;
            min-height: 100vh;
            display: flex;
            align-items: center;
            justify-content: center;
            padding: 1.5rem;
        }

        .container {
            width: 100%;
            max-width: 780px;
        }

        /* Header */
        .header {
            text-align: center;
            margin-bottom: 2rem;
        }
        .header .eyebrow {
            font-size: 0.7rem;
            font-weight: 600;
            letter-spacing: 0.18em;
            text-transform: uppercase;
            color: #6366f1;
            margin-bottom: 0.6rem;
        }
        .header h1 {
            font-size: 1.75rem;
            font-weight: 700;
            color: #111827;
            letter-spacing: -0.03em;
            margin-bottom: 0.3rem;
        }
        .header p {
            font-size: 0.875rem;
            color: #6b7280;
        }

        /* Grid layout */
        .grid {
            display: grid;
            grid-template-columns: 1fr 1fr;
            gap: 1rem;
            margin-bottom: 1rem;
        }

        /* Panels */
        .panel {
            background: #ffffff;
            border: 1px solid #e5e7eb;
            border-radius: 12px;
            padding: 1.25rem;
            box-shadow: 0 1px 4px rgba(0,0,0,0.06);
        }

        .panel-label {
            font-size: 0.7rem;
            font-weight: 600;
            letter-spacing: 0.12em;
            text-transform: uppercase;
            color: #9ca3af;
            margin-bottom: 1rem;
            display: flex;
            align-items: center;
            gap: 0.4rem;
        }

        /* Date input */
        input[type="date"] {
            width: 100%;
            padding: 0.65rem 0.85rem;
            background: #f9fafb;
            border: 1px solid #e5e7eb;
            border-radius: 8px;
            color: #111827;
            font-family: 'Inter', sans-serif;
            font-size: 0.9rem;
            outline: none;
            transition: border-color 0.2s, box-shadow 0.2s;
            cursor: pointer;
        }
        input[type="date"]::-webkit-calendar-picker-indicator {
            filter: none;
            cursor: pointer;
            opacity: 0.5;
        }
        input[type="date"]:focus {
            border-color: #6366f1;
            box-shadow: 0 0 0 3px rgba(99,102,241,0.12);
        }

        /* City cards */
        .cities-grid {
            display: grid;
            grid-template-columns: repeat(3, 1fr);
            gap: 0.6rem;
        }

        .city-card {
            background: #f9fafb;
            border: 1px solid #e5e7eb;
            border-radius: 8px;
            padding: 0.7rem 0.85rem;
            cursor: pointer;
            transition: border-color 0.2s, background 0.2s, box-shadow 0.2s;
            user-select: none;
            pointer-events: all;
        }
        .city-card * { pointer-events: none; }

        .city-card:hover {
            border-color: #c7d2fe;
            background: #f0f1ff;
        }
        .city-card.active {
            border-color: #6366f1;
            background: #eef2ff;
            box-shadow: 0 0 0 3px rgba(99,102,241,0.1);
        }

        .city-row {
            display: flex;
            align-items: center;
            justify-content: space-between;
        }
        .city-name {
            font-size: 0.85rem;
            font-weight: 600;
            color: #111827;
        }
        .city-code {
            font-size: 0.7rem;
            font-weight: 700;
            color: #d1d5db;
            letter-spacing: 0.05em;
        }
        .city-card.active .city-code { color: #a5b4fc; }

        .city-badge {
            display: inline-block;
            font-size: 0.65rem;
            font-weight: 500;
            padding: 0.15rem 0.45rem;
            border-radius: 4px;
            margin-top: 0.35rem;
        }
        .badge-active  { background: #dcfce7; color: #16a34a; }
        .badge-config  { background: #fef9c3; color: #ca8a04; }
        .badge-beta    { background: #e0e7ff; color: #4f46e5; }
        .badge-all     { background: #f3e8ff; color: #7c3aed; }

        /* Check dot */
        .check {
            width: 14px;
            height: 14px;
            border-radius: 50%;
            border: 1.5px solid #d1d5db;
            margin-top: 0.35rem;
            transition: all 0.2s;
            display: flex;
            align-items: center;
            justify-content: center;
        }
        .city-card.active .check {
            background: #6366f1;
            border-color: #6366f1;
        }
        .check::after {
            content: '';
            width: 5px;
            height: 5px;
            border-radius: 50%;
            background: white;
            opacity: 0;
            transition: opacity 0.2s;
        }
        .city-card.active .check::after { opacity: 1; }

        /* Button */
        .btn {
            width: 100%;
            padding: 0.8rem 1rem;
            background: #6366f1;
            border: none;
            border-radius: 10px;
            color: #fff;
            font-family: 'Inter', sans-serif;
            font-size: 0.9rem;
            font-weight: 600;
            cursor: pointer;
            display: flex;
            align-items: center;
            justify-content: center;
            gap: 0.5rem;
            transition: background 0.2s, transform 0.1s, box-shadow 0.2s;
            margin-top: 1rem;
            box-shadow: 0 2px 8px rgba(99,102,241,0.3);
        }
        .btn:hover { background: #4f52d6; box-shadow: 0 4px 14px rgba(99,102,241,0.4); }
        .btn:active { transform: scale(0.99); }
        .btn:disabled { background: #e5e7eb; color: #9ca3af; cursor: not-allowed; transform: none; box-shadow: none; }

        .spinner {
            width: 16px; height: 16px;
            border: 2px solid rgba(255,255,255,0.4);
            border-top-color: #fff;
            border-radius: 50%;
            animation: spin 0.7s linear infinite;
            display: none;
        }
        @keyframes spin { to { transform: rotate(360deg); } }

        /* Toasts */
        .toasts {
            position: fixed;
            bottom: 1.5rem;
            right: 1.5rem;
            display: flex;
            flex-direction: column;
            gap: 0.5rem;
            z-index: 999;
        }
        .toast {
            background: #ffffff;
            border: 1px solid #e5e7eb;
            border-radius: 10px;
            padding: 0.85rem 1rem;
            min-width: 270px;
            max-width: 340px;
            display: flex;
            align-items: flex-start;
            gap: 0.7rem;
            box-shadow: 0 8px 24px rgba(0,0,0,0.1);
            transform: translateX(120%);
            transition: transform 0.3s cubic-bezier(0.16,1,0.3,1);
        }
        .toast.show { transform: translateX(0); }
        .toast-ok  { border-left: 3px solid #22c55e; }
        .toast-err { border-left: 3px solid #ef4444; }
        .t-icon { font-size: 0.95rem; margin-top: 1px; }
        .t-body { flex: 1; }
        .t-title { font-size: 0.82rem; font-weight: 600; color: #111827; }
        .t-msg   { font-size: 0.76rem; color: #6b7280; margin-top: 0.1rem; word-break: break-all; }
        .t-close {
            background: none; border: none; color: #9ca3af;
            cursor: pointer; font-size: 1rem; line-height: 1; padding: 0;
        }
        .t-close:hover { color: #111827; }
    </style>
    <div class="container">

        <div class="header">
            <div class="eyebrow">XD Allocation Hub</div>
            <h1>Generate Report</h1>
            <p>Select the PO date and city to compile and download</p>
        </div>

        <div class="grid">
            <!-- Date Panel -->
            <div class="panel">
                <div class="panel-label">
                    <svg width="12" height="12" viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2.5"><rect x="3" y="4" width="18" height="18" rx="2"/><line x1="16" y1="2" x2="16" y2="6"/><line x1="8" y1="2" x2="8" y2="6"/><line x1="3" y1="10" x2="21" y2="10"/></svg>
                    Date (D-1 / PO Date)
                </div>
                <input type="date" id="date-picker">
            </div>

            <!-- City Panel -->
            <div class="panel">
                <div class="panel-label">
                    <svg width="12" height="12" viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2.5"><path d="M12 22s-8-4.5-8-11.8A8 8 0 0 1 12 2a8 8 0 0 1 8 8.2c0 7.3-8 11.8-8 11.8z"/><circle cx="12" cy="10" r="3"/></svg>
                    Select City
                </div>
                <div class="cities-grid">
                    <div class="city-card active" data-city="chennai">
                        <div class="city-row">
                            <span class="city-name">Chennai</span>
                            <span class="city-code">MAA</span>
                        </div>
                        <div class="city-row" style="margin-top:0.35rem">
                            <span class="city-badge badge-active">Active</span>
                            <div class="check"></div>
                        </div>
                    </div>
                    <div class="city-card" data-city="trichy">
                        <div class="city-row">
                            <span class="city-name">Trichy</span>
                            <span class="city-code">TRZ</span>
                        </div>
                        <div class="city-row" style="margin-top:0.35rem">
                            <span class="city-badge badge-active">Active</span>
                            <div class="check"></div>
                        </div>
                    </div>
                    <div class="city-card" data-city="coimbatore">
                        <div class="city-row">
                            <span class="city-name">Coimbatore</span>
                            <span class="city-code">CJB</span>
                        </div>
                        <div class="city-row" style="margin-top:0.35rem">
                            <span class="city-badge badge-active">Active</span>
                            <div class="check"></div>
                        </div>
                    </div>
                    <div class="city-card" data-city="bengaluru">
                        <div class="city-row">
                            <span class="city-name">Bengaluru</span>
                            <span class="city-code">BLR</span>
                        </div>
                        <div class="city-row" style="margin-top:0.35rem">
                            <span class="city-badge badge-active">Active</span>
                            <div class="check"></div>
                        </div>
                    </div>
                    <div class="city-card" data-city="mumbai">
                        <div class="city-row">
                            <span class="city-name">Mumbai</span>
                            <span class="city-code">BOM</span>
                        </div>
                        <div class="city-row" style="margin-top:0.35rem">
                            <span class="city-badge badge-active">Active</span>
                            <div class="check"></div>
                        </div>
                    </div>
                    <div class="city-card" data-city="all">
                        <div class="city-row">
                            <span class="city-name">All Cities</span>
                            <span class="city-code">ALL</span>
                        </div>
                        <div class="city-row" style="margin-top:0.35rem">
                            <span class="city-badge badge-all">ZIP</span>
                            <div class="check"></div>
                        </div>
                    </div>
                </div>
            </div>
        </div>

        <input type="hidden" id="selected-city" value="chennai">

        <button class="btn" id="btn-compile" onclick="compile()">
            <svg id="btn-icon" width="15" height="15" viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2.5" stroke-linecap="round" stroke-linejoin="round">
                <path d="M21 15v4a2 2 0 0 1-2 2H5a2 2 0 0 1-2-2v-4"/>
                <polyline points="7 10 12 15 17 10"/>
                <line x1="12" y1="15" x2="12" y2="3"/>
            </svg>
            <div class="spinner" id="btn-spinner"></div>
            <span id="btn-label">Compile &amp; Download Report</span>
        </button>
    </div>

    <div class="toasts" id="toasts"></div>

    <script>
        // Pre-fill today's date
        const dp = document.getElementById('date-picker');
        const now = new Date();
        dp.value = `${now.getFullYear()}-${String(now.getMonth()+1).padStart(2,'0')}-${String(now.getDate()).padStart(2,'0')}`;

        // City card selection
        document.querySelectorAll('.city-card').forEach(card => {
            card.addEventListener('click', function() {
                document.querySelectorAll('.city-card').forEach(c => c.classList.remove('active'));
                this.classList.add('active');
                document.getElementById('selected-city').value = this.dataset.city;
            });
        });

        async function compile() {
            const date = dp.value;
            const city = document.getElementById('selected-city').value;
            const btn  = document.getElementById('btn-compile');
            const icon = document.getElementById('btn-icon');
            const spin = document.getElementById('btn-spinner');
            const lbl  = document.getElementById('btn-label');

            if (!date) { showToast('Missing date', 'Please select a date first.', 'err'); return; }

            btn.disabled = true;
            icon.style.display = 'none';
            spin.style.display = 'block';
            lbl.textContent = 'Processing...';

            try {
                const res = await fetch('/process', {
                    method: 'POST',
                    headers: { 'Content-Type': 'application/json' },
                    body: JSON.stringify({ date, city })
                });

                if (!res.ok) {
                    let msg = 'Request failed.';
                    try { msg = (await res.json()).error || msg; } catch(_) {}
                    throw new Error(msg);
                }

                // Parse filename from Content-Disposition header
                const disp = res.headers.get('Content-Disposition') || '';
                const match = disp.match(/filename[^;=\\n]*=(['"]*)(.*?)\\1(?:;|$)/);
                const filename = match ? match[2].trim() :
                    (city === 'all' ? `XD_All_${date}.zip` : `XD_${city}_${date}.xlsx`);

                const blob = await res.blob();
                const url  = URL.createObjectURL(blob);
                const a    = Object.assign(document.createElement('a'), { href: url, download: filename });
                document.body.appendChild(a);
                a.click();
                a.remove();
                URL.revokeObjectURL(url);

                showToast('Downloaded', filename, 'ok');
            } catch(e) {
                showToast('Error', e.message, 'err');
            } finally {
                btn.disabled = false;
                icon.style.display = '';
                spin.style.display = 'none';
                lbl.textContent = 'Compile & Download Report';
            }
        }

        function showToast(title, msg, type) {
            const box = document.getElementById('toasts');
            const el  = document.createElement('div');
            el.className = `toast toast-${type}`;
            el.innerHTML = `
                <span class="t-icon">${type === 'ok' ? '✓' : '✕'}</span>
                <div class="t-body">
                    <div class="t-title">${title}</div>
                    <div class="t-msg">${msg}</div>
                </div>
                <button class="t-close" onclick="this.closest('.toast').remove()">×</button>
            `;
            box.appendChild(el);
            requestAnimationFrame(() => requestAnimationFrame(() => el.classList.add('show')));
            setTimeout(() => { el.classList.remove('show'); setTimeout(() => el.remove(), 350); }, 5000);
        }
    </script>
</body>
</html>
"""

# ─────────────────────────────────────────────
# FLASK API
# ─────────────────────────────────────────────
def create_app():
    import io, zipfile
    from flask import Flask, request, jsonify, send_file
    app = Flask(__name__)

    @app.route("/", methods=["GET"])
    def index():
        return INDEX_HTML

    @app.route("/health", methods=["GET"])
    def health():
        return jsonify({"status": "ok", "cities": list(CITIES.keys())})

    @app.route("/process", methods=["POST"])
    def process():
        body = request.get_json(force=True, silent=True) or {}
        date_str = body.get("date") or request.args.get("date")
        city_key = (body.get("city") or request.args.get("city", "all")).lower()

        if not date_str:
            return jsonify({"error": "Missing 'date' field. Send { \"date\": \"YYYY-MM-DD\" }"}), 400

        if city_key not in list(CITIES.keys()) + ["all"]:
            return jsonify({"error": f"Unknown city '{city_key}'. Valid: {list(CITIES.keys())} or 'all'"}), 400

        try:
            target_date = datetime.strptime(date_str, "%Y-%m-%d").date()
        except ValueError:
            return jsonify({"error": "Invalid date format. Use YYYY-MM-DD"}), 400

        delivery_date = target_date + timedelta(days=1)
        # Format month as short title-case name (e.g., Jun_27)
        delivery_str  = delivery_date.strftime('%b_%d')

        cities_to_run = list(CITIES.keys()) if city_key == "all" else [city_key]

        if len(cities_to_run) == 1:
            # ── Single city – return xlsx directly ──────────────────────
            ck    = cities_to_run[0]
            label = CITIES[ck]["label"]
            fname = f"{label}_XD_Allocation_{delivery_str}_Delivery.xlsx"
            opath = os.path.join(OUTPUT_DIR, fname)
            try:
                run_city_report(target_date, opath, ck)
            except Exception as e:
                import traceback
                traceback.print_exc()
                return jsonify({"error": str(e)}), 500
            return send_file(
                opath,
                as_attachment=True,
                download_name=fname,
                mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        else:
            # ── All cities – return a ZIP containing all xlsx files ──────
            errors = {}
            zip_buf = io.BytesIO()
            with zipfile.ZipFile(zip_buf, "w", zipfile.ZIP_DEFLATED) as zf:
                for ck in cities_to_run:
                    label = CITIES[ck]["label"]
                    fname = f"{label}_XD_Allocation_{delivery_str}_Delivery.xlsx"
                    opath = os.path.join(OUTPUT_DIR, fname)
                    try:
                        run_city_report(target_date, opath, ck)
                        zf.write(opath, fname)
                    except Exception as e:
                        errors[ck] = str(e)
            zip_buf.seek(0)
            if errors and len(errors) == len(cities_to_run):
                return jsonify({"error": "All cities failed", "details": errors}), 500
            zip_name = f"XD_Allocation_All_{delivery_str}Delivery.zip"
            resp = send_file(
                zip_buf,
                as_attachment=True,
                download_name=zip_name,
                mimetype="application/zip",
            )
            if errors:
                resp.headers["X-Partial-Errors"] = json.dumps(errors)
            return resp

    return app


# ─────────────────────────────────────────────
# CLI ENTRY POINT
# ─────────────────────────────────────────────
if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="XD Allocation Processor – Multi-City")
    subparsers = parser.add_subparsers(dest="cmd")

    # run as CLI
    run_parser = subparsers.add_parser("run", help="Process a specific date")
    run_parser.add_argument("--date", required=True, help="Date in YYYY-MM-DD format")
    run_parser.add_argument(
        "--city",
        default="all",
        choices=list(CITIES.keys()) + ["all"],
        help="City to process (default: all)",
    )
    run_parser.add_argument("--out", default=None, help="Output file path (only used when --city is a single city)")

    # run as Flask server
    serve_parser = subparsers.add_parser("serve", help="Start Flask API server")
    serve_parser.add_argument("--port", default=5050, type=int)

    args = parser.parse_args()

    if args.cmd == "run":
        target_date   = datetime.strptime(args.date, "%Y-%m-%d").date()
        delivery_date = target_date + timedelta(days=1)
        delivery_str  = delivery_date.strftime('%b_%d')

        cities_to_run = list(CITIES.keys()) if args.city == "all" else [args.city]

        for city_key in cities_to_run:
            label = CITIES[city_key]["label"]
            if args.out and len(cities_to_run) == 1:
                out = args.out
            else:
                out = os.path.join(OUTPUT_DIR, f"{label}_XD_Allocation_{delivery_str}_Delivery.xlsx")
            try:
                run_city_report(target_date, out, city_key)
            except RuntimeError as e:
                print(f"[SKIP] {label}: {e}")

    elif args.cmd == "serve":
        app = create_app()
        print(f"[OK] XD Processor API running on http://0.0.0.0:{args.port}")
        app.run(host="0.0.0.0", port=args.port, debug=False)

    else:
        parser.print_help()
